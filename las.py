
    # import os

    # os.chdir("/Users/gonghongwei/Desktop/openpyxl")
import numpy as np
# 这个包要包装
from openpyxl import load_workbook
# from openpyxl import Workbook

# import openpyxl
import xml.etree.ElementTree as ET

FileName1 = 'test1.xlsx'
FileName2 = 'test2.xlsx'
SheetNameS = [u'Sheet1', u'Sheet2']
#SheetName = u'Sheet1'


# 替换函数开始
def replace_xls(filename1, filename2, sheetname):
    wb = load_workbook(filename1)
    wb2 = load_workbook(filename2)
    ws = wb[sheetname]
    ws2 = wb2[sheetname]
    # 两个for循环遍历整个excel的单元格内容
    for i, row in enumerate(ws.iter_rows()):
        for j, cell in enumerate(row):
            ws2.cell(row=i + 1, column=j + 1, value=cell.value)
    wb2.save(filename2)


def replace_xls(filename1, filename2, sheetname1, sheetname2):
    wb = load_workbook(filename1)
    wb2 = load_workbook(filename2)
    ws = wb[sheetname1]
    ws2 = wb2[sheetname2]
    # 两个for循环遍历整个excel的单元格内容
    for i, row in enumerate(ws.iter_rows()):
        for j, cell in enumerate(row):
            ws2.cell(row=i + 1, column=j + 1, value=cell.value)
    wb2.save(filename2)


def directory_replace_1(directory, row):
    # 这就是一个2019年后清华同方版本的目录更替
    print('第一种目录更替，这就是一个2019年后清华同方版本的目录更替', directory, row)


if __name__ == '__main__':
    in_file = open('data.xml')
    tree = ET.parse(in_file)
    root = tree.getroot()  # 获取根节点
    i = 0
    for child in root:

        print(root[i][0].text)
        print(root[i][1].text)
        
        i = i+1
        # 开始单行判断 进行相应的，目录更替
        # 如果,第二列是1，进入第一个 但参数要是两个，一个是目录（原行的第一列），别一个是第几行。因为要依据第几行来判断命名的表名。
        # 第一行的（也就是序号为0）表名是当期，上期，同期】
        Directory = '一个目录'
        # directory_replace_1(Directory, i)
        # 如果，第二个是2，进入第二个


#    for SheetName in SheetNameS:
#        replace_xls(FileName1, FileName2, SheetName)
#    replace_xls(FileName1, FileName2, SheetName)


